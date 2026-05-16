# -*- coding: utf-8 -*-
"""
RAG Engine - Handles file parsing, embedding, and retrieval for the chatbot.
Supports: PDF, CSV, TXT, Images (OCR)
"""

import sys
import io
# Force UTF-8 output encoding for Windows compatibility
if sys.stdout.encoding != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
if sys.stderr.encoding != 'utf-8':
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

import os
import csv
import hashlib
import json
import re
import gc
from pathlib import Path
from typing import Optional
from dotenv import load_dotenv
# ใช้ google-genai (SDK ใหม่) แทน google-generativeai (deprecated)
from google import genai as _genai_client_factory
import time
from datetime import datetime

try:
    import fitz # PyMuPDF
    PDF_SUPPORTED = True
except ImportError:
    PDF_SUPPORTED = False

import chromadb
from chromadb.utils import embedding_functions

# Load environment variables early and override any existing session variables
BASE_DIR = Path(__file__).parent.absolute()
load_dotenv(BASE_DIR / ".env", override=True)

try:
    import docx
    DOCX_SUPPORTED = True
except ImportError:
    DOCX_SUPPORTED = False

try:
    import openpyxl
    XLSX_SUPPORTED = True
except ImportError:
    XLSX_SUPPORTED = False

try:
    from PIL import Image
    import pytesseract
    IMAGE_SUPPORTED = True
except ImportError:
    IMAGE_SUPPORTED = False


BASE_DIR = Path(__file__).parent.absolute()
UPLOAD_DIR = BASE_DIR / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

CHROMA_DIR = BASE_DIR / "chroma_db"
CHROMA_DIR.mkdir(exist_ok=True)

META_FILE = BASE_DIR / "file_meta.json"


def _load_meta() -> dict:
    if META_FILE.exists():
        with open(META_FILE, "r", encoding="utf-8") as f:
            meta = json.load(f)
            
        # --- Path Self-Healing for Migration (Windows <-> Pi/Linux) ---
        modified = False
        for fid, info in meta.items():
            stored_path = info.get("path")
            if stored_path:
                # If the path looks like Windows (has \) but we are on Linux (or vice versa)
                # Or if the parent directory no longer matches our actual UPLOAD_DIR
                # Use regex to split by both / and \ to safely get the filename on any OS
                # Use a robust way to extract filename from either Windows or Linux absolute paths
                actual_name = stored_path.replace('\\', '/').split('/')[-1]
                expected_path = str(UPLOAD_DIR / actual_name)
                
                if stored_path != expected_path:
                    # Fix it to match current OS and movement of folder
                    info["path"] = expected_path
                    modified = True
                    
        if modified:
            print(f"[HEALED] Paths in file_meta.json self-healed. Total reconciled: {len([m for m in meta if meta[m].get('path')])}")
            _save_meta(meta)
            
        return meta
    return {}


def _save_meta(meta: dict):
    with open(META_FILE, "w", encoding="utf-8") as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)


def _file_hash(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def _chunk_text(text: str, chunk_size: int = 800, overlap: int = 150) -> list[str]:
    """Split text into overlapping chunks, optimized for Thai and English."""
    # Pre-clean: collapse whitespace but keep some structure
    text = re.sub(r"([ \t])+", " ", text).strip()
    
    chunks = []
    start = 0
    text_len = len(text)
    
    # Common separators for Thai and English
    # In Thai, spaces are used as phrase/sentence delimiters
    # Thai markers: ๆ (repeat), ฯ (omission), ๆลๆ (etc)
    # Thai conjunctions: ซึ่ง, และ, หรือ, เพราะ, ดังนั้น, ที่
    separators = [
        "\n\n", "\n", 
        ". ", "! ", "? ", 
        " ๆลๆ", " ฯลฯ", " ๆ ", " ฯ ", 
        " ซึ่ง", " และ", " หรือ", " เพราะ", " ดังนั้น", " โดยเฉพาะ", " อย่างไรก็ตาม",
        "  ", ". ", " "
    ] 
    
    while start < text_len:
        end = min(start + chunk_size, text_len)
        
        # Try to find a logical break point within the last 25% of the chunk
        if end < text_len:
            lookback_limit = max(start, end - 200)
            best_break = -1
            for sep in separators:
                idx = text.rfind(sep, lookback_limit, end)
                if idx != -1:
                    best_break = idx + len(sep)
                    break
            
            if best_break != -1:
                end = best_break
        
        chunk = text[start:end].strip()
        if chunk and len(chunk) > 5:  # Avoid tiny fragments
            chunks.append(chunk)
            
        # Advance with overlap
        new_start = end - overlap
        if new_start <= start:
            start = end 
        else:
            # Ensure overlap doesn't start in the middle of a short fragment
            start = new_start
            
    return chunks


# ─────────────────────────────────────────────
# Vector Store
# ─────────────────────────────────────────────
class BatchedGoogleEmbeddingFunction(embedding_functions.EmbeddingFunction):
    """ฟังก์ชัน embedding โดยใช้ google-genai SDK ใหม่ แทน deprecated GoogleGenerativeAi."""
    def __init__(self, api_key: str, model_name: str = "models/gemini-embedding-001"):
        self._api_key = api_key
        self._model_name = model_name
        self._client = _genai_client_factory.Client(api_key=api_key, http_options={"api_version": "v1beta"})
        self.daily_quota_hit = False

    def __call__(self, input: list[str]) -> list[list[float]]:
        if not input: return []
        batch_size = 100
        all_embeddings = []
        for i in range(0, len(input), batch_size):
            batch = input[i : i + batch_size]
            try:
                # ใช้ google-genai Client API ใหม่
                response = self._client.models.embed_content(
                    model=self._model_name,
                    contents=batch,
                    config={"task_type": "retrieval_document"}
                )
                # SDK ใหม่คืน response.embeddings เป็น list[ContentEmbedding]
                for emb in response.embeddings:
                    all_embeddings.append(emb.values)
                if len(input) > batch_size:
                    time.sleep(5.0)
            except Exception as e:
                err_msg = str(e).lower()
                if "429" in err_msg or "quota" in err_msg:
                    if any(x in err_msg for x in ["per day", "daily", "limit: 1000", "perday"]):
                        print("🚫 DAILY Quota Exceeded. หยุดการ embed ชั่วคราว")
                        self.daily_quota_hit = True
                        raise Exception("DAILY_QUOTA_EXCEEDED")
                    print(f"⚠️ Rate Limit — รอ 30 วินาทีแล้วลองใหม่...")
                    time.sleep(30.0)
                    try:
                        response = self._client.models.embed_content(
                            model=self._model_name,
                            contents=batch,
                            config={"task_type": "retrieval_document"}
                        )
                        for emb in response.embeddings:
                            all_embeddings.append(emb.values)
                        continue
                    except:
                        pass
                print(f"❌ Embedding API Error: {e}")
                raise e
        return all_embeddings

class OllamaEmbeddingFunction(embedding_functions.EmbeddingFunction):
    """Custom embedding function using Ollama."""
    def __init__(self, model_name="mxbai-embed-large"):
        self.model_name = model_name
        self.url = "http://localhost:11434/api/embeddings"

    def __call__(self, input: list[str]) -> list[list[float]]:
        if not input: return []
        all_embeddings = []
        for text in input:
            try:
                response = requests.post(
                    self.url,
                    json={"model": self.model_name, "prompt": text}
                )
                if response.status_code == 200:
                    all_embeddings.append(response.json()["embedding"])
                else:
                    raise Exception(f"Ollama Error: {response.text}")
            except Exception as e:
                print(f"❌ Ollama Embedding Error: {e}")
                raise e
        return all_embeddings

class FastEmbedEmbeddingFunction(embedding_functions.EmbeddingFunction):
    """Truly local embedding function using FastEmbed."""
    def __init__(self, model_name="sentence-transformers/all-MiniLM-L6-v2"):
        try:
            from fastembed import TextEmbedding
            # Define a local cache directory to avoid Windows Temp permission issues
            local_cache = os.path.join(os.getcwd(), "fastembed_cache")
            if not os.path.exists(local_cache):
                os.makedirs(local_cache, exist_ok=True)
            self.model = TextEmbedding(model_name=model_name, cache_dir=local_cache)
        except ImportError:
            print("❌ 'fastembed' not installed. Please run: pip install fastembed")
            raise ImportError("fastembed not installed")

    def __call__(self, input: list[str]) -> list[list[float]]:
        if not input: return []
        # FastEmbed returns a generator of numpy arrays
        embeddings = list(self.model.embed(input))
        return [l.tolist() for l in embeddings]

import requests # Ensure requests is available in rag_engine

class KnowledgeBase:
    def __init__(self):
        # Use absolute path for ChromaDB
        CHROMA_DIR.mkdir(parents=True, exist_ok=True)
        
        self.api_key = os.environ.get("GEMINI_API_KEY", "").strip()
        if not self.api_key:
            print("⚠️ Warning: GEMINI_API_KEY not found in environment.")
            self.api_key = "dummy_key_for_init"
        self.provider = os.environ.get("AI_EMBEDDING_PROVIDER", "gemini").strip().lower()
        print(f"DEBUG: AI_EMBEDDING_PROVIDER from env is: '{self.provider}'")
        provider = self.provider
        if provider == "ollama":
            embed_model = os.environ.get("OLLAMA_EMBED_MODEL", "mxbai-embed-large")
            print(f"🧠 Using Ollama Embeddings (Model: {embed_model})")
            ef = OllamaEmbeddingFunction(model_name=embed_model)
        elif provider == "local":
            print(f"[LOCAL] Using Local FastEmbed Embeddings")
            try:
                ef = FastEmbedEmbeddingFunction()
            except ImportError:
                print("⚠️ Falling back to Dummy embeddings because fastembed is missing.")
                # We'll throw an error later if they try to use it
                ef = None
        else:
            print(f"☁️ ใช้ Google Gemini Embeddings (google-genai SDK ใหม่)")
            ef = BatchedGoogleEmbeddingFunction(
                api_key=self.api_key,
                model_name="models/gemini-embedding-001"
            )
        
        if ef is None:
             raise Exception("EMBEDDING_PROVIDER_NOT_READY")
        try:
            client = chromadb.PersistentClient(path=str(CHROMA_DIR))
            self.collection = client.get_or_create_collection(
                name="org_knowledge",
                embedding_function=ef,
                metadata={"hnsw:space": "cosine"},
            )
        except Exception as e:
            # Self-healing for ChromaDB version/type mismatches
            print(f"⚠️ ChromaDB Error detected: {e}. Attempting self-healing...")
            import shutil
            if CHROMA_DIR.exists():
                shutil.rmtree(CHROMA_DIR)
            CHROMA_DIR.mkdir(exist_ok=True)
            # Wipe meta too to stay in sync
            if META_FILE.exists():
                os.remove(META_FILE)
            
            client = chromadb.PersistentClient(path=str(CHROMA_DIR))
            self.collection = client.get_or_create_collection(
                name="org_knowledge",
                embedding_function=ef,
                metadata={"hnsw:space": "cosine"},
            )
            print("✅ Database reset and ready.")

    def update_api_key(self, api_key: str):
        """อัปเดต API key และรีเซ็ต collection"""
        if not api_key or api_key == "dummy_key_for_init":
            return
        self.api_key = api_key
        ef = BatchedGoogleEmbeddingFunction(
            api_key=api_key,
            model_name="models/gemini-embedding-001"
        )
        client = chromadb.PersistentClient(path=str(CHROMA_DIR))
        self.collection = client.get_or_create_collection(
            name="org_knowledge",
            embedding_function=ef,
            metadata={"hnsw:space": "cosine"},
        )
        print("✅ KnowledgeBase API Key อัปเดตแล้ว collection รีเซ็ตเรียบร้อย")

    def is_key_valid(self):
        """Simple check if we have a real key if provider is Gemini."""
        if self.provider != "gemini":
            return True # Local providers don't need GEMINI_API_KEY
        return self.api_key and self.api_key != "dummy_key_for_init"

    def add_chunks(self, chunks: list[str], source_name: str, file_id: str, metadatas: list[dict] = None):
        ids = [f"{file_id}_chunk_{i}" for i in range(len(chunks))]
        if metadatas is None:
            metadatas = [{"source": source_name, "file_id": file_id} for _ in chunks]
        else:
            # Merge with default metadata
            for m in metadatas:
                m["source"] = source_name
                m["file_id"] = file_id
        
        try:
            self.collection.upsert(documents=chunks, ids=ids, metadatas=metadatas)
        except Exception as e:
            err_msg = str(e).lower()
            # Only recreate if it's a GENUINE dimension mismatch (e.g. model changed)
            # Standard Chroma dimension error contains 'Expected' and 'got'
            if "dimension" in err_msg and ("expected" in err_msg or "but got" in err_msg):
                print("⚠️ Critical Dimension mismatch detected. Recreating collection for new model...")
                # To change dimensions, we MUST delete the collection and recreate it.
                client = chromadb.PersistentClient(path=str(CHROMA_DIR))
                try:
                    client.delete_collection("org_knowledge")
                except:
                    pass
                
                # Re-init collection using current embedding function
                self.collection = client.get_or_create_collection(
                    name="org_knowledge",
                    embedding_function=self.collection._embedding_function,
                    metadata={"hnsw:space": "cosine"},
                )
                # Retry once
                self.collection.upsert(documents=chunks, ids=ids, metadatas=metadatas)
            else:
                raise e

    def query(self, question: str, n_results: int = 4, where: dict = None) -> list[dict]:
        if not self.is_key_valid():
            print("⚠️ API Key not valid. Skipping query.")
            return []
            
        count = self.collection.count()
        if count == 0:
            return []
        n = min(n_results, count)
        result = self.collection.query(query_texts=[question], n_results=n, where=where)
        docs = result.get("documents", [[]])[0]
        metas = result.get("metadatas", [[]])[0]
        distances = result.get("distances", [[]])[0]
        
        results = []
        for d, m, dist in zip(docs, metas, distances):
            # Distance 0 is perfect. Normalize to a 0-1 score
            score = max(0, 1 - (dist / 1.5))
            item = {"text": d, "score": score}
            item.update(m)
            results.append(item)
        return results

    def delete_by_file_id(self, file_id: str):
        try:
            # Direct deletion by metadata filter is more robust in recent Chroma versions
            self.collection.delete(where={"file_id": file_id})
        except Exception as e:
            print(f"⚠️ ChromaDB deletion warning: {e}")
            # Fallback: ingest might overwrite anyway, so we continue

    def delete_by_source(self, source_name: str):
        """Used for deleting specific Wiki pages or similar by source string."""
        try:
            self.collection.delete(where={"source": source_name})
        except Exception as e:
            print(f"⚠️ ChromaDB source deletion warning: {e}")

    def update_metadata_by_file_id(self, file_id: str, new_metas_dict: dict):
        """Updates metadata for all chunks with this file_id."""
        try:
            results = self.collection.get(where={"file_id": file_id})
            ids = results.get("ids", [])
            if not ids:
                return
            
            existing_metas = results.get("metadatas", [])
            updated_metas = []
            for meta in existing_metas:
                # Normalize new metadata: convert None to "", others to str
                for k, v in new_metas_dict.items():
                    if k == "category_id":
                        meta[k] = str(v) if v is not None else ""
                    else:
                        meta[k] = v
                updated_metas.append(meta)
            
            self.collection.update(ids=ids, metadatas=updated_metas)
        except Exception as e:
            print(f"⚠️ ChromaDB update error for {file_id}: {e}")

    def fix_chroma_categories(self, meta: dict):
        """One-time utility to sync category_ids from meta file to ChromaDB chunks."""
        print("🛠️ Syncing missing category_ids to ChromaDB...")
        for fid, info in meta.items():
            cat = info.get("category_id")
            # Sync ALL states, normalizing None to ""
            val = str(cat) if cat is not None else ""
            self.update_metadata_by_file_id(fid, {"category_id": val})
        print("✅ Sync complete.")

    def total_chunks(self) -> int:
        return self.collection.count()

    def hybrid_query(self, question: str, n_results: int = 5, where: dict = None) -> list[dict]:
        """Combines vector search with restricted keyword-based matching for Pi efficiency."""
        if not self.is_key_valid():
            return []
            
        count = self.collection.count()
        if count == 0:
            return []

        # 1. Semantic Search (Vector) - Fetch double the requested results
        # This is where the heavy lifting happens via Gemini Cloud (Fast)
        vector_results = self.query(question, n_results=n_results * 2, where=where)
        
        # 2. Optimized Keyword Match
        # We only check keywords within the vector results to avoid RAM overflow
        keywords = [kw.lower() for kw in question.split() if len(kw) > 1]
        if not keywords:
            return vector_results[:n_results]
        
        for item in vector_results:
            lower_text = item["text"].lower()
            matches = [kw for kw in keywords if kw in lower_text]
            if matches:
                # Boost score if keywords found
                item["score"] = min(1.0, item["score"] + (len(matches) / len(keywords)) * 0.3)
        
        # 3. Final Sort and Trim
        final_list = sorted(vector_results, key=lambda x: x["score"], reverse=True)
        return final_list[:n_results]

    def prune_orphaned_chunks(self, valid_file_ids: list[str]) -> int:
        """Remove chunks that don't belong to any file in the current metadata."""
        try:
            # This is a bit expensive as we have to fetch all IDs or iterate
            all_data = self.collection.get(include=['metadatas'])
            all_metas = all_data.get('metadatas', [])
            all_ids = all_data.get('ids', [])
            
            to_delete = []
            for i, meta in enumerate(all_metas):
                fid = meta.get('file_id')
                if fid not in valid_file_ids:
                    to_delete.append(all_ids[i])
            
            if to_delete:
                # Delete in batches to avoid overwhelming the system
                batch_size = 100
                for i in range(0, len(to_delete), batch_size):
                    self.collection.delete(ids=to_delete[i:i+batch_size])
                return len(to_delete)
            return 0
        except Exception as e:
            print(f"⚠️ Pruning error: {e}")
            return 0

    def wipe(self):
        """Clear all data from the collection."""
        ids = self.collection.get()["ids"]
        if ids:
            self.collection.delete(ids=ids)


# File Processor
# ─────────────────────────────────────────────
class FileProcessor:
    @staticmethod
    def parse(file_path: Path) -> tuple[list[str], list[dict]]:
        """Returns (chunks, metadatas)."""
        suffix = file_path.suffix.lower()
        if suffix == ".pdf":
            return FileProcessor._parse_pdf(file_path)
        elif suffix == ".csv":
            return FileProcessor._parse_csv(file_path)
        elif suffix == ".docx":
            return FileProcessor._parse_docx(file_path)
        elif suffix in (".xlsx", ".xls"):
            return FileProcessor._parse_xlsx(file_path)
        elif suffix in (".txt", ".md", ".json"):
            return FileProcessor._parse_text(file_path)
        elif suffix in (".png", ".jpg", ".jpeg", ".bmp", ".tiff", ".webp"):
            return FileProcessor._parse_image(file_path)
        else:
            return [], []

    @staticmethod
    def _parse_pdf(path: Path) -> tuple[list[str], list[dict]]:
        if not PDF_SUPPORTED:
            return ["[PDF parsing unavailable]"], [{"location": "error"}]
        chunks, metas = [], []
        doc = fitz.open(str(path))
        for i, page in enumerate(doc, 1):
            text = page.get_text()
            page_chunks = _chunk_text(text)
            for chunk in page_chunks:
                chunks.append(chunk)
                metas.append({"location": f"Page {i}"})
        doc.close()
        return chunks, metas

    @staticmethod
    def _parse_csv(path: Path) -> tuple[list[str], list[dict]]:
        """Parse CSV and group rows to reduce chunk count. Includes Thai encoding fallback."""
        chunks, metas = [], []
        # Try multiple encodings for Thai CSV compatibility
        encodings = ["utf-8-sig", "utf-8", "cp874", "tis-620"]
        
        success = False
        for enc in encodings:
            try:
                with open(path, newline="", encoding=enc) as f:
                    # Use a sample to detect if it's actually readable with this encoding
                    sample = f.read(1024)
                    f.seek(0)
                    reader = csv.DictReader(f)
                    rows = []
                    i = 0
                    for i, row in enumerate(reader, 1):
                        row_text = ", ".join(f"{k}:{v}" for k, v in row.items() if v)
                        if row_text.strip():
                            rows.append(row_text)
                        
                        # Group every 15 rows into one chunk (increased for efficiency)
                        if len(rows) >= 15:
                            chunks.append("\n".join(rows))
                            metas.append({"location": f"Rows {i-14}-{i}"})
                            rows = []
                        
                        # Safety cap for Pi memory: max 500 chunks per file (Optimized for RAM)
                        if len(chunks) >= 500:
                            print(f"⚠️ File too large for Pi RAM. Capping at 500 chunks.")
                            break
                    
                    # Remaining rows
                    if rows:
                        chunks.append("\n".join(rows))
                        metas.append({"location": f"Rows {i-len(rows)+1}-{i}"})
                
                success = True
                print(f"✅ CSV parsed successfully with {enc} encoding.")
                break
            except (UnicodeDecodeError, UnicodeError):
                continue
            except Exception as e:
                print(f"⚠️ CSV parsing error with {enc}: {e}")
                continue
        
        if not success:
            return [f"[CSV Error: Could not decode file with supported encodings]"], [{"location": "error"}]
            
        return chunks, metas

    @staticmethod
    def _parse_docx(path: Path) -> tuple[list[str], list[dict]]:
        if not DOCX_SUPPORTED:
            return ["[DOCX parsing unavailable]"], [{"location": "error"}]
        try:
            doc = docx.Document(path)
            full_text = "\n".join([para.text for para in doc.paragraphs])
            chunks = _chunk_text(full_text)
            return chunks, [{"location": "Document"}] * len(chunks)
        except Exception as e:
            return [f"[DOCX Error: {e}]"], [{"location": "error"}]

    @staticmethod
    def _parse_xlsx(path: Path) -> tuple[list[str], list[dict]]:
        if not XLSX_SUPPORTED:
            return ["[XLSX parsing unavailable]"], [{"location": "error"}]
        chunks, metas = [], []
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for i, row in enumerate(ws.iter_rows(values_only=True), 1):
                    row_text = " | ".join(str(cell) for cell in row if cell is not None)
                    if row_text.strip():
                        chunks.append(row_text)
                        metas.append({"location": f"Sheet: {sheet}, Row {i}"})
            return chunks, metas
        except Exception as e:
            return [f"[Excel Error: {e}]"], [{"location": "error"}]

    @staticmethod
    def _parse_text(path: Path) -> tuple[list[str], list[dict]]:
        try:
            text = path.read_text(encoding="utf-8")
        except Exception:
            text = path.read_text(encoding="cp874", errors="replace")
        chunks = _chunk_text(text)
        return chunks, [{"location": "Document"}] * len(chunks)

    @staticmethod
    def _parse_image(path: Path) -> tuple[list[str], list[dict]]:
        if not IMAGE_SUPPORTED:
            return ["[OCR unavailable]"], [{"location": "error"}]
        try:
            img = Image.open(str(path))
            text = pytesseract.image_to_string(img, lang="tha+eng")
            chunks = _chunk_text(text) or ["[No text found in image]"]
            return chunks, [{"location": "Image OCR"}] * len(chunks)
        except Exception as e:
            return [f"[OCR Error: {e}]"], [{"location": "error"}]


# ─────────────────────────────────────────────
# Public API used by app.py
# ─────────────────────────────────────────────
_kb = KnowledgeBase()

# Re-init in case load_dotenv was called in app.py after this module was first loaded
def reinit_kb():
    global _kb
    print("🔄 Re-initializing KnowledgeBase to refresh settings...")
    _kb = KnowledgeBase()
    # Sync missing category IDs to ChromaDB one-time on startup
    try:
        meta = _load_meta()
        _kb.fix_chroma_categories(meta)
    except Exception as e:
        print(f"⚠️ Initial Chroma Sync error: {e}")


def ingest_file(file_path: Path, original_name: str, department: str = "General", category_id: int = None) -> dict:
    """Parse file, chunk, embed, and store in ChromaDB. Returns metadata dict."""
    meta = _load_meta()
    fid = _file_hash(file_path)

    print(f"[INGEST] Ingesting: {original_name} ({fid}) | Dept: {department}")
    if fid in meta and meta[fid].get("status") == "ready":
        print(f"⏩ Duplicate file: {original_name}")
        return {"status": "duplicate", "file_id": fid, "name": original_name}

    # Initial state to show in UI while processing
    meta[fid] = {
        "name": original_name,
        "path": str(file_path),
        "chunks": 0,
        "size": file_path.stat().st_size,
        "type": file_path.suffix.lower().lstrip("."),
        "department": department,
        "category_id": category_id,
        "status": "processing",
        "api_key_status": _kb.api_key[:5] + "..." if _kb.is_key_valid() else "MISSING KEY",
        "timestamp": datetime.now().isoformat()
    }
    _save_meta(meta)

    if not _kb.is_key_valid():
        print(f"❌ API Key not valid while ingesting {original_name}")
        return {"status": "error", "error": "API Key not set. Cannot ingest file.", "file_id": fid, "name": original_name}

    try:
        chunks, chunk_metas = FileProcessor.parse(file_path)
    except Exception as e:
        print(f"❌ Parsing Error: {e}")
        return {"status": "error", "error": f"Parsing failed: {e}", "file_id": fid, "name": original_name}
    if not chunks:
        return {"status": "empty", "file_id": fid, "name": original_name}

    for m in chunk_metas:
        m["department"] = department
        # Always set category_id as a string (use "" if not provided)
        m["category_id"] = str(category_id) if category_id is not None else ""

    try:
        # Step-by-step ingestion to save RAM
        batch_size = 20 # Smaller batches for Pi
        for i in range(0, len(chunks), batch_size):
            batch_chunks = chunks[i : i + batch_size]
            batch_metas = chunk_metas[i : i + batch_size]
            _kb.add_chunks(batch_chunks, source_name=original_name, file_id=fid, metadatas=batch_metas)
            gc.collect() 
    except Exception as e:
        if "DAILY_QUOTA_EXCEEDED" in str(e):
             raise e # Propagate hard stop
        print(f"❌ Error adding chunks for {original_name}: {e}")
        return {"status": "error", "error": str(e), "file_id": fid, "name": original_name}
    finally:
        gc.collect() # Force cleanup for Pi memory

    meta[fid] = {
        "name": original_name,
        "path": str(file_path),
        "chunks": len(chunks),
        "size": file_path.stat().st_size,
        "type": file_path.suffix.lower().lstrip("."),
        "department": department,
        "category_id": category_id,
        "status": "ready",
        "api_key_status": _kb.api_key[:5] + "..." if _kb.is_key_valid() else "MISSING KEY",
        "timestamp": meta[fid].get("timestamp", datetime.now().isoformat())
    }
    _save_meta(meta)
    return {"status": "ok", "file_id": fid, "name": original_name, "chunks": len(chunks), "department": department, "category_id": category_id}


def ingest_text(text: str, source_name: str, department: str = "General", category_id: int = None) -> dict:
    """Useful for ingesting Wiki pages or AI-generated content directly."""
    if not text.strip():
        return {"status": "empty"}
    
    # Create a deterministic ID for source
    import hashlib
    source_id = hashlib.md5(source_name.encode()).hexdigest()[:12]
    
    # First, clear previous version of this source if it exists
    _kb.delete_by_source(source_name)
    
    chunks = _chunk_text(text)
    chunk_metas = []
    for _ in chunks:
        m = {"department": department, "source": source_name, "type": "wiki"}
        m["category_id"] = category_id # Always set
        chunk_metas.append(m)

    try:
        _kb.add_chunks(chunks, source_name=source_name, file_id=f"text_{source_id}", metadatas=chunk_metas)
        return {"status": "ok", "chunks": len(chunks), "source": source_name}
    except Exception as e:
        print(f"❌ ingest_text Error: {e}")
        return {"status": "error", "error": str(e)}


def list_files() -> list[dict]:
    meta = _load_meta()
    return [{"file_id": fid, **info} for fid, info in meta.items()]


def delete_file(file_id: str, delete_from_disk: bool = True) -> bool:
    meta = _load_meta()
    if file_id not in meta:
        return False
    _kb.delete_by_file_id(file_id)
    info = meta.pop(file_id)
    _save_meta(meta)
    if delete_from_disk:
        try:
            Path(info["path"]).unlink(missing_ok=True)
        except Exception:
            pass
    return True

def update_file_category(file_id: str, category_id: int) -> bool:
    """Updates the category assigned to a file in metadata."""
    meta = _load_meta()
    if file_id not in meta:
        return False
    meta[file_id]["category_id"] = category_id
    _save_meta(meta)
    
    # Also update vector database chunks
    _kb.update_metadata_by_file_id(file_id, {"category_id": category_id})
    return True


def get_file_content(file_id: str) -> str:
    """Retrieve all text chunks for a specific file_id from vector store."""
    try:
        results = _kb.collection.get(where={"file_id": file_id})
        docs = results.get("documents", [])
        if not docs:
            # Fallback: check if it's a 'source' name instead of hash ID
            results = _kb.collection.get(where={"source": file_id})
            docs = results.get("documents", [])
        return "\n".join(docs)
    except Exception as e:
        print(f"❌ Error getting file content for {file_id}: {e}")
        return ""


def retrieve_context(question: str, where: dict = None) -> tuple[str, list[dict]]:
    """Return (context_text, list_of_detailed_sources) for the question."""
    print(f"[SEARCH] Starting RAG retrieval for: '{question[:20]}...' Filter: {where}")
    final_results = []
    
    try:
        # 1. Filename-based boost (Optimized for Thai: No split needed)
        filename_chunks = []
        if question and len(question) > 2:
            try:
                # 1a. Efficiently get candidates: Fetch unique source names first
                inv_data = _kb.collection.get(include=['metadatas'], limit=100)
                sources_in_db = list(set([m.get("source") for m in inv_data.get("metadatas", []) if m.get("source")]))
                
                # 1b. Match filenames against the RAW question (Better for Thai)
                matched_sources = [s for s in sources_in_db if s.lower() in question.lower()]
                
                if matched_sources:
                    print(f"📂 Thai Match filenames: {matched_sources}")
                    # Fetch chunks for matched files
                    f_res = _kb.collection.get(
                        where={"source": {"$in": matched_sources}} if not where else {"$and": [where, {"source": {"$in": matched_sources}}]},
                        limit=5
                    )
                    if f_res and f_res['documents']:
                        for i in range(len(f_res['documents'])):
                            filename_chunks.append({
                                "text": f_res['documents'][i],
                                "score": 1.0,
                                **f_res['metadatas'][i]
                            })
            except Exception as fe:
                print(f"⚠️ Filename match error: {fe}")
        
        # If it's a very short question or casual greeting, skip expensive RAG
        # Added: Better detection for Thai number gibberish or pure noise
        clean_q = re.sub(r'[0-9๐-๙\s\W]', '', question)
        if not question or len(question) < 5 or (len(clean_q) < 2 and len(question) < 15):
            print(f"ℹ️ Fast Path: Skipping RAG for noisy/short query: '{question}'", flush=True)
            return "", []
        
        # 2. Standard Hybrid Query (Semantic + Keyword)
        results = _kb.hybrid_query(question, where=where)
        # Filter by score (keep only >= 0.42) - Increased from 0.35 to reduce noise
        results = [r for r in results if r.get("score", 0) >= 0.42]
        
        # Merge results, prioritize filename chunks
        seen_texts = set()
        for r in filename_chunks + results:
            text_small = r.get("text", "")[:200]
            if text_small not in seen_texts:
                final_results.append(r)
                seen_texts.add(text_small)
                
        print(f"✅ RAG query finished. Returned {len(final_results)} relevant chunks.")
    except Exception as e:
        print(f"❌ RAG query error: {e}")
        return "", []

    if not final_results:
        print("ℹ️ No relevant context found.")
        return "", []
    
    results = final_results[:10] # Limit context size
    
    # Enrich context with location info
    context_parts = []
    detailed_sources = []
    seen_source_keys = set()
    
    for r in results:
        loc = r.get("location", "Document")
        dept = r.get("department", "General")
        src = r.get("source", "?")
        fid = r.get("file_id", "")
        
        # Build context text for AI
        context_parts.append(f"--- แหล่งที่มา: {src} [{loc}] ({dept}) ---\n{r['text']}")
        
        # Build structured sources for UI
        source_key = f"{fid}_{loc}"
        if source_key not in seen_source_keys:
            detailed_sources.append({
                "file_id": fid,
                "name": src,
                "location": loc,
                "department": dept,
                "type": r.get("type", "file")
            })
            seen_source_keys.add(source_key)
        
    context = "\n\n".join(context_parts)
    print(f"[CONTEXT] Context prepared: {len(context)} chars, {len(detailed_sources)} detailed sources.")
    return context, detailed_sources


def kb_stats() -> dict:
    return {"total_chunks": _kb.total_chunks(), "total_files": len(_load_meta())}


def get_quota_status() -> dict:
    """Check quota for current embedding function."""
    res = {"embedding_quota_hit": False, "provider": "gemini"}
    ef = getattr(_kb.collection, "_embedding_function", None)
    if ef:
        if isinstance(ef, BatchedGoogleEmbeddingFunction):
            res["embedding_quota_hit"] = ef.daily_quota_hit
            res["provider"] = "gemini"
        elif isinstance(ef, OllamaEmbeddingFunction):
            res["embedding_quota_hit"] = False # Local doesn't have daily quota
            res["provider"] = "ollama"
    return res


def prune_kb() -> int:
    """Consolidate and remove orphaned chunks from the vector database."""
    meta = _load_meta()
    valid_ids = list(meta.keys())
    return _kb.prune_orphaned_chunks(valid_ids)

def wipe_knowledge_base():
    """Wipe all data from the knowledge base and clear meta."""
    meta = _load_meta()
    for fid, info in meta.items():
        try:
            Path(info["path"]).unlink(missing_ok=True)
        except Exception:
            pass
def fix_categories():
    """Sync and normalize category_ids from meta file to ChromaDB."""
    meta = _load_meta()
    return _kb.fix_chroma_categories(meta)

def sync_uploads():
    """Scan the uploads folder and ingest any files that are not already in meta."""
    meta = _load_meta()
    existing_hashes = set(meta.keys())
    results = []
    
    if not UPLOAD_DIR.exists():
        return results

    # Removed is_key_valid check as we use local embeddings now

    for file_path in UPLOAD_DIR.iterdir():
        if file_path.is_file():
            try:
                fid = _file_hash(file_path)
                
                # DEEP SYNC FIX: Check if the file actually has data in ChromaDB
                in_db = False
                try:
                    db_check = _kb.collection.get(where={"file_id": fid}, limit=1)
                    if db_check and db_check.get("ids"):
                        in_db = True
                except:
                    pass

                if fid not in existing_hashes or not in_db:
                    print(f"[SYNC] Re-ingesting or Syncing file: {file_path.name}")
                    original_name = file_path.name
                    if len(file_path.name) > 9 and file_path.name[8] == "_":
                        original_name = file_path.name[9:]
                    
                    res = ingest_file(file_path, original_name=original_name)
                    results.append(res)
            except Exception as e:
                print(f"⚠️ Error syncing {file_path.name}: {e}")
                continue
    return results


def search_kb(question: str, n_results: int = 5, where: dict = None) -> list[dict]:
    """Perform a raw search in the knowledge base with optional permissions filter."""
    return _kb.hybrid_query(question, n_results=n_results, where=where)

def query(question: str, n_results: int = 4, where: dict = None) -> list[dict]:
    """Alias for search_kb to maintain compatibility."""
    return search_kb(question, n_results=n_results, where=where)

def get_all_chunks(where: dict = None, limit: int = 20) -> list[dict]:
    """Fetch raw chunks from the knowledge base without semantic search."""
    data = _kb.collection.get(where=where, limit=limit)
    docs = data.get("documents", [])
    metas = data.get("metadatas", [])
    
    # Simple deduplication by content
    seen = set()
    results = []
    for d, m in zip(docs, metas):
        if d in seen: continue
        seen.add(d)
        item = {"text": d, "score": 0.5}
        item.update(m)
        results.append(item)
    return results
