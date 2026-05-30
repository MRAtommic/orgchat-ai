import openpyxl
import json
import os

file_path = r"c:\Users\KC_Ketwilai\Downloads\orgchat-ai-main\orgchat-ai-main\webai\ภงด.53\peak\PEAK_ImportExpense.xlsx"
out_path = r"c:\Users\KC_Ketwilai\Downloads\orgchat-ai-main\orgchat-ai-main\webai\peak_info.json"

res = {}
if os.path.exists(file_path):
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        res["sheets"] = wb.sheetnames
        res["data"] = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for r in range(1, 6):
                row_vals = [ws.cell(r, c).value for c in range(1, min(ws.max_column + 1, 40))]
                if any(row_vals is not None for row_vals in row_vals):
                    rows.append(row_vals)
            res["data"][sheet_name] = rows
        res["success"] = True
    except Exception as e:
        res["success"] = False
        res["error"] = str(e)
else:
    res["success"] = False
    res["error"] = "File not found"

with open(out_path, "w", encoding="utf-8") as f:
    json.dump(res, f, ensure_ascii=False, indent=2)

print("Peak info extracted successfully!")
